"""
export.py — Excel Export
Generates a .xlsx workbook with:
  - Summary
  - Cost Breakdown
  - Revenue
  - Cash Flow (120-month waterfall)
  - Assumptions & Sources
  - Zoning Adjustments
"""

import io
from datetime import date
from typing import Optional
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from calculations import build_cash_flow_waterfall

# Color palette
GREEN_FILL  = PatternFill("solid", fgColor="C6EFCE")
RED_FILL    = PatternFill("solid", fgColor="FFC7CE")
BLUE_FILL   = PatternFill("solid", fgColor="BDD7EE")
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
SUBHEAD_FILL = PatternFill("solid", fgColor="2E75B6")
GRAY_FILL   = PatternFill("solid", fgColor="F2F2F2")

WHITE_FONT  = Font(color="FFFFFF", bold=True)
BOLD_FONT   = Font(bold=True)
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

MONEY_FMT   = '"$"#,##0'
PCT_FMT     = '0.00%'
NUM_FMT     = '#,##0'
FLOAT2_FMT  = '#,##0.00'


def _set_col_width(ws, col_letter: str, width: float):
    ws.column_dimensions[col_letter].width = width


def _header_row(ws, row: int, values: list, fill=HEADER_FILL):
    for i, val in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=val)
        c.font = HEADER_FONT
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN_BORDER


def _label_value(ws, row: int, label: str, value, fmt: str = None, indent: int = 0):
    lc = ws.cell(row=row, column=1, value=("  " * indent) + label)
    lc.font = Font(bold=(indent == 0))
    lc.border = THIN_BORDER

    vc = ws.cell(row=row, column=2, value=value)
    if fmt:
        vc.number_format = fmt
    vc.alignment = Alignment(horizontal="right")
    vc.border = THIN_BORDER
    return vc


def _section_header(ws, row: int, label: str, ncols: int = 2):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=label)
    c.font = WHITE_FONT
    c.fill = SUBHEAD_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")


# -----------------------------------------------------------------------
# Tab 1: Summary
# -----------------------------------------------------------------------
def _build_summary(wb, results: dict, user_inputs: dict):
    ws = wb.create_sheet("Summary")
    _set_col_width(ws, "A", 40)
    _set_col_width(ws, "B", 22)

    # Title
    ws.merge_cells("A1:B1")
    t = ws["A1"]
    t.value = "Real Estate Development Feasibility — Summary"
    t.font = Font(bold=True, size=14, color="FFFFFF")
    t.fill = HEADER_FILL
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Feasibility badge
    ws.merge_cells("A2:B2")
    badge = ws["A2"]
    is_feasible = results.get("is_feasible", False)
    badge.value = "FEASIBLE ✓" if is_feasible else "NOT FEASIBLE ✗"
    badge.font = Font(bold=True, size=16, color="FFFFFF")
    badge.fill = PatternFill("solid", fgColor="375623") if is_feasible else PatternFill("solid", fgColor="9C0006")
    badge.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 40

    row = 4
    _section_header(ws, row, "Project Overview"); row += 1
    _label_value(ws, row, "Location", user_inputs.get("location", "")); row += 1
    _label_value(ws, row, "Use Type", user_inputs.get("use_type", "")); row += 1
    _label_value(ws, row, "Building Type", user_inputs.get("building_type", "")); row += 1
    _label_value(ws, row, "Total Units", results.get("num_units", 0), NUM_FMT); row += 1
    _label_value(ws, row, "Parcel Size (acres)", user_inputs.get("parcel_acres", 0), FLOAT2_FMT); row += 1
    _label_value(ws, row, "Report Date", date.today().isoformat()); row += 2

    _section_header(ws, row, "Key Financial Metrics"); row += 1
    _label_value(ws, row, "Total Development Cost", results.get("total_dev_cost", 0), MONEY_FMT); row += 1
    _label_value(ws, row, "Cost per Unit", results.get("cost_per_unit", 0), MONEY_FMT); row += 1
    _label_value(ws, row, "NOI (Stabilized)", results.get("noi", 0), MONEY_FMT); row += 1
    _label_value(ws, row, "Return on Cost (Actual)", results.get("return_on_cost", 0), PCT_FMT); row += 1
    _label_value(ws, row, "Return on Cost (Threshold)", results.get("return_on_cost_threshold", 0.06), PCT_FMT); row += 1
    if results.get("for_sale_margin") is not None:
        _label_value(ws, row, "Profit Margin (For-Sale)", results.get("for_sale_margin", 0), PCT_FMT); row += 1
    _label_value(ws, row, "Exit Value (Cap Rate)", results.get("exit_value", 0), MONEY_FMT); row += 1
    if results.get("irr") is not None:
        _label_value(ws, row, "Levered IRR (5yr)", results.get("irr", 0), PCT_FMT); row += 1
    row += 1

    _section_header(ws, row, "Feasibility Analysis"); row += 1
    ws.cell(row=row, column=1, value=results.get("verdict_explanation", "")).font = Font(italic=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 3, end_column=2)
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    row += 4

    # Color the return on cost cell
    roc_row = None
    for r in range(1, row):
        c = ws.cell(row=r, column=1)
        if c.value and "Return on Cost (Actual)" in str(c.value):
            roc_row = r
            break
    if roc_row:
        vc = ws.cell(row=roc_row, column=2)
        vc.fill = GREEN_FILL if is_feasible else RED_FILL


# -----------------------------------------------------------------------
# Tab 2: Cost Breakdown
# -----------------------------------------------------------------------
def _build_costs(wb, results: dict):
    ws = wb.create_sheet("Cost Breakdown")
    _set_col_width(ws, "A", 38)
    _set_col_width(ws, "B", 20)
    _set_col_width(ws, "C", 20)

    _header_row(ws, 1, ["Cost Category", "Total ($)", "Per Unit ($)"])
    rows = [
        ("Land",                 results.get("land_cost", 0)),
        ("Hard Costs (Vertical)",results.get("hard_costs", 0)),
        ("Parking Construction", results.get("parking_hard_cost", 0)),
        ("Soft Costs (18%)",     results.get("soft_costs", 0)),
        ("Construction Interest",results.get("construction_interest", 0)),
    ]
    num_units = results.get("num_units", 1) or 1

    for i, (label, total) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=label).border = THIN_BORDER
        vc = ws.cell(row=i, column=2, value=total)
        vc.number_format = MONEY_FMT
        vc.border = THIN_BORDER
        vc.alignment = Alignment(horizontal="right")
        puc = ws.cell(row=i, column=3, value=total / num_units)
        puc.number_format = MONEY_FMT
        puc.border = THIN_BORDER
        puc.alignment = Alignment(horizontal="right")
        if i % 2 == 0:
            for col in range(1, 4):
                ws.cell(row=i, column=col).fill = GRAY_FILL

    total_row = len(rows) + 2
    ws.cell(row=total_row, column=1, value="TOTAL DEVELOPMENT COST").font = BOLD_FONT
    ws.cell(row=total_row, column=1).border = THIN_BORDER
    tc = ws.cell(row=total_row, column=2, value=results.get("total_dev_cost", 0))
    tc.number_format = MONEY_FMT
    tc.font = BOLD_FONT
    tc.fill = BLUE_FILL
    tc.border = THIN_BORDER
    tc.alignment = Alignment(horizontal="right")
    puc = ws.cell(row=total_row, column=3, value=results.get("cost_per_unit", 0))
    puc.number_format = MONEY_FMT
    puc.font = BOLD_FONT
    puc.fill = BLUE_FILL
    puc.border = THIN_BORDER
    puc.alignment = Alignment(horizontal="right")

    # Key metrics section
    r = total_row + 2
    ws.cell(row=r, column=1, value="Key Sizing Assumptions").font = BOLD_FONT
    r += 1
    sizing = [
        ("Total NSF", f"{results.get('total_nsf', 0):,.0f} SF"),
        ("Total GSF", f"{results.get('total_gsf', 0):,.0f} SF"),
        ("Net-to-Gross Ratio", f"{results.get('ntg_ratio', 0):.0%}"),
        ("Weighted Avg Unit Size", f"{results.get('weighted_avg_unit_size_sf', 0):,.0f} SF"),
        ("Parking Spaces", f"{results.get('num_parking_spaces', 0):,}"),
        ("Hard Cost / GSF", f"${results.get('hard_cost_per_gsf', 0):,.0f}"),
        ("Land Cost / SF", f"${results.get('land_cost_per_sf', 0):,.2f}"),
    ]
    for label, val in sizing:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=val).alignment = Alignment(horizontal="right")
        r += 1


# -----------------------------------------------------------------------
# Tab 3: Revenue
# -----------------------------------------------------------------------
def _build_revenue(wb, results: dict):
    ws = wb.create_sheet("Revenue")
    _set_col_width(ws, "A", 38)
    _set_col_width(ws, "B", 22)

    _header_row(ws, 1, ["Revenue / Expense Item", "Annual ($)"])

    rows_data = [
        ("REVENUE", None, SUBHEAD_FILL),
        ("Market Rent Revenue",     results.get("market_rent_annual", 0), None),
        ("Affordable Rent Revenue", results.get("affordable_rent_annual", 0), None),
        ("Parking Revenue",         results.get("parking_revenue_annual", 0), None),
        ("Gross Revenue",           results.get("gross_revenue", 0), BLUE_FILL),
        ("Less: Vacancy",           -results.get("gross_revenue", 0) * results.get("vacancy_rate", 0.05), None),
        ("Effective Gross Income (EGI)", results.get("egi", 0), BLUE_FILL),
        ("EXPENSES", None, SUBHEAD_FILL),
        ("Operating Expenses",      -results.get("total_opex", 0), None),
        ("Management Fee",          -results.get("mgmt_fee", 0), None),
        ("Property Taxes",          -results.get("property_taxes", 0), None),
        ("CapEx Reserve",           -results.get("capex_reserve", 0), None),
        ("Total Expenses",          -results.get("total_expenses", 0), RED_FILL),
        ("NET OPERATING INCOME",    results.get("noi", 0),
            GREEN_FILL if results.get("noi", 0) >= 0 else RED_FILL),
    ]

    for i, row_item in enumerate(rows_data, start=2):
        label, value, fill = row_item
        lc = ws.cell(row=i, column=1, value=label)
        lc.border = THIN_BORDER
        if fill:
            lc.fill = fill
        if label in ("REVENUE", "EXPENSES"):
            lc.font = WHITE_FONT
        elif label in ("Gross Revenue", "Effective Gross Income (EGI)", "Total Expenses", "NET OPERATING INCOME"):
            lc.font = BOLD_FONT

        if value is not None:
            vc = ws.cell(row=i, column=2, value=value)
            vc.number_format = MONEY_FMT
            vc.alignment = Alignment(horizontal="right")
            vc.border = THIN_BORDER
            if fill:
                vc.fill = fill
            if label in ("Gross Revenue", "Effective Gross Income (EGI)", "Total Expenses", "NET OPERATING INCOME"):
                vc.font = BOLD_FONT

    # Key metrics
    r = len(rows_data) + 3
    ws.cell(row=r, column=1, value="Revenue Metrics").font = BOLD_FONT
    r += 1
    metrics = [
        ("Vacancy Rate",             f"{results.get('vacancy_rate', 0):.1%}"),
        ("Return on Cost",           f"{results.get('return_on_cost', 0):.2%}"),
        ("Cap Rate",                 f"{results.get('cap_rate', 0):.2%}"),
        ("Exit Value",               f"${results.get('exit_value', 0):,.0f}"),
        ("Weighted Monthly Rent",    f"${results.get('weighted_monthly_rent', 0):,.0f}"),
        ("Required Monthly Rent",    f"${results.get('required_monthly_rent', 0):,.0f}"),
    ]
    for label, val in metrics:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=val).alignment = Alignment(horizontal="right")
        r += 1


# -----------------------------------------------------------------------
# Tab 4: Cash Flow (120-month)
# -----------------------------------------------------------------------
def _build_cash_flow(wb, results: dict):
    ws = wb.create_sheet("Cash Flow (120-Month)")
    headers = ["Month", "Period", "Gross Revenue", "Vacancy Loss", "EGI",
               "OpEx", "NOI", "Debt Service", "Levered CF"]
    _header_row(ws, 1, headers)
    widths = [8, 14, 16, 16, 16, 16, 16, 16, 16]
    for i, w in enumerate(widths, start=1):
        _set_col_width(ws, get_column_letter(i), w)

    waterfall = build_cash_flow_waterfall(results)
    for i, row in enumerate(waterfall, start=2):
        vals = [
            row["month"], row["period"], row["gross_revenue"], row["vacancy_loss"],
            row["egi"], row["opex"], row["noi"], row["debt_service"], row["levered_cf"],
        ]
        bg = GRAY_FILL if i % 2 == 0 else None
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="right" if j > 2 else "left")
            if bg:
                c.fill = bg
            if j > 2:
                c.number_format = MONEY_FMT


# -----------------------------------------------------------------------
# Tab 5: Assumptions & Sources
# -----------------------------------------------------------------------
def _build_assumptions(wb, assumptions: dict):
    ws = wb.create_sheet("Assumptions & Sources")
    headers = ["Category", "Assumption", "Value", "Unit", "Source Name", "Source URL", "Date Retrieved", "Notes"]
    _header_row(ws, 1, headers)
    widths = [18, 30, 16, 14, 20, 40, 14, 40]
    for i, w in enumerate(widths, start=1):
        _set_col_width(ws, get_column_letter(i), w)

    row = 2
    for category, cat_data in assumptions.items():
        if not isinstance(cat_data, dict):
            continue
        for key, item in cat_data.items():
            if not isinstance(item, dict):
                continue
            c = ws.cell(row=row, column=1, value=category)
            c.border = THIN_BORDER
            ws.cell(row=row, column=2, value=key).border = THIN_BORDER
            vc = ws.cell(row=row, column=3, value=item.get("value"))
            vc.border = THIN_BORDER
            vc.alignment = Alignment(horizontal="right")
            ws.cell(row=row, column=4, value=item.get("unit")).border = THIN_BORDER
            ws.cell(row=row, column=5, value=item.get("source_name")).border = THIN_BORDER
            url_cell = ws.cell(row=row, column=6, value=item.get("source_url"))
            url_cell.border = THIN_BORDER
            url_cell.alignment = Alignment(wrap_text=False)
            ws.cell(row=row, column=7, value=item.get("date_retrieved")).border = THIN_BORDER
            notes_cell = ws.cell(row=row, column=8, value=item.get("notes"))
            notes_cell.border = THIN_BORDER
            notes_cell.alignment = Alignment(wrap_text=True)
            if row % 2 == 0:
                for col in range(1, 9):
                    ws.cell(row=row, column=col).fill = GRAY_FILL
            row += 1


# -----------------------------------------------------------------------
# Tab 6: Zoning Adjustments
# -----------------------------------------------------------------------
def _build_zoning_adjustments(wb, zoning_result: dict):
    ws = wb.create_sheet("Zoning Adjustments")
    headers = ["Adjustment Type", "Description", "Original Value", "Revised Value",
               "Unit", "Source URL", "Confidence", "ROC Impact", "Notes"]
    _header_row(ws, 1, headers)
    widths = [20, 40, 14, 14, 12, 40, 12, 12, 40]
    for i, w in enumerate(widths, start=1):
        _set_col_width(ws, get_column_letter(i), w)

    applicable = zoning_result.get("applicable_adjustments", [])
    all_adjustments = zoning_result.get("adjustments", [])

    if not all_adjustments:
        ws.cell(row=2, column=1, value="No zoning adjustments found for this location.")
        return

    for i, adj in enumerate(all_adjustments, start=2):
        conf = adj.get("confidence", "low")
        is_applied = adj in applicable

        vals = [
            adj.get("adjustment_type", ""),
            adj.get("description", ""),
            adj.get("original_value"),
            adj.get("revised_value"),
            adj.get("unit", ""),
            adj.get("source_url", ""),
            conf,
            adj.get("roc_impact"),
            adj.get("notes", ""),
        ]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.border = THIN_BORDER
            if j == 8 and v is not None:
                c.number_format = PCT_FMT
                c.fill = GREEN_FILL if (v or 0) >= 0 else RED_FILL
            if not is_applied:
                c.font = Font(color="A0A0A0", italic=True)  # Gray out non-applied
            c.alignment = Alignment(wrap_text=(j in (2, 6, 9)), horizontal="left")

    # Summary row
    r = len(all_adjustments) + 3
    ws.cell(row=r, column=1, value="Summary").font = BOLD_FONT
    ws.cell(row=r + 1, column=1, value=f"Adjustments found: {len(all_adjustments)}")
    ws.cell(row=r + 2, column=1, value=f"Adjustments applied (high/med confidence): {len(applicable)}")
    ws.cell(row=r + 3, column=1, value=f"Total ROC impact: {zoning_result.get('roc_delta', 0):.2%}")


# -----------------------------------------------------------------------
# Main export function
# -----------------------------------------------------------------------
def export_to_excel(
    results: dict,
    user_inputs: dict,
    assumptions: dict,
    zoning_result: dict,
) -> bytes:
    """
    Build the workbook and return as bytes for Streamlit download.
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    _build_summary(wb, results, user_inputs)
    _build_costs(wb, results)
    _build_revenue(wb, results)
    _build_cash_flow(wb, results)
    _build_assumptions(wb, assumptions)
    _build_zoning_adjustments(wb, zoning_result)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
